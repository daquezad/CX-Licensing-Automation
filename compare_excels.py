import sys
import os
import argparse
import logging
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from utils.date_utils import standardize_date
from utils.colors import RED_FILL, BLUE_FILL, YELLOW_FILL, GREEN_FILL
from utils.logging_utils import setup_logging
from utils.mapping_utils import load_pid_to_skus_map, get_valid_sku_matches
from utils.fs_utils import ensure_clean_dir
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def main(pre_ea_path, cssm_path, pid_to_skus_map: dict[str, list[str]]):
    logger = logging.getLogger(__name__)
    logger.info("Starting comparison: pre_ea=%s cssm=%s", pre_ea_path, cssm_path)
    start_time = datetime.now()

    # Read CSSM as DataFrame
    logger.info("Reading CSSM 'License Detail' sheet starting at row 6: %s", cssm_path)
    cssm = pd.read_excel(cssm_path, sheet_name='License Detail', header=5)
    cssm.columns = cssm.columns.map(str.strip)
    cssm['Source Identifier'] = cssm['Source Identifier'].astype(str).str.strip()
    cssm['SKU'] = cssm['SKU'].astype(str).str.strip()

    # Output directory (created/cleaned in __main__ before logging is configured)
    output_dir = os.path.join(os.getcwd(), "output_files")

    # Copy pre-ea.xlsx to output_files/pre-ea_compared.xlsx using openpyxl (preserve all content)
    base_name = os.path.basename(pre_ea_path)
    out_name = base_name.replace('.xlsx', '_compared.xlsx')
    out_path = os.path.join(output_dir, out_name)
    wb = load_workbook(pre_ea_path)
    wb.save(out_path)
    wb = load_workbook(out_path)
    ws = wb.active

    # Read PRE-EA as DataFrame for easy row iteration
    pre_ea = pd.read_excel(pre_ea_path)
    pre_ea.columns = pre_ea.columns.map(str.strip)

    logger.info("Loaded PRE-EA rows: %d | CSSM rows: %d", len(pre_ea), len(cssm))

    red_rows, blue_rows, yellow_rows, green_rows = 0, 0, 0, 0

    for idx, row in pre_ea.iterrows():
        alc_order_number = row.get('ALC Order Number')
        pre_ea_migrated_pid = row.get('Pre EA Migrated Pid')
        pre_ea_qty = row.get('Quantity')
        pre_ea_exp = row.get('Expiration Date')
        excel_row_idx = idx + 2  # for Excel row index (header + 0-index)

        alc_order_number_str = str(alc_order_number).strip()
        pre_ea_migrated_pid_str = str(pre_ea_migrated_pid).strip()

        # Find all rows in CSSM where Source Identifier matches
        cssm_matches = cssm[cssm['Source Identifier'] == alc_order_number_str]

        if cssm_matches.empty:
            logger.info("Row %d: ALC Order Number '%s' NOT found in CSSM. Marking as RED.",
                        excel_row_idx, alc_order_number_str)
            for col in range(1, len(pre_ea.columns) + 1):
                ws.cell(row=excel_row_idx, column=col).fill = RED_FILL
            red_rows += 1
            continue

        # --- Use the external mapping for exceptions here ---
        sku_match = get_valid_sku_matches(cssm_matches, pre_ea_migrated_pid_str, pid_to_skus_map)

        if sku_match.empty:
            logger.info("Row %d: No SKU '%s' (or mapped exception) for ALC Order Number '%s' in CSSM. Marking as RED.",
                        excel_row_idx, pre_ea_migrated_pid_str, alc_order_number_str)
            for col in range(1, len(pre_ea.columns) + 1):
                ws.cell(row=excel_row_idx, column=col).fill = RED_FILL
            red_rows += 1
            continue

        cssm_row = sku_match.iloc[0]
        cssm_qty = cssm_row['Available To Use']

        try:
            cssm_qty = int(cssm_qty)
        except Exception:
            cssm_qty = None

        if cssm_qty != pre_ea_qty:
            logger.info("Row %d: Quantity mismatch (PRE-EA: %s, CSSM: %s). Marking as BLUE.",
                        excel_row_idx, pre_ea_qty, cssm_qty)
            for col in range(1, len(pre_ea.columns) + 1):
                ws.cell(row=excel_row_idx, column=col).fill = BLUE_FILL
            blue_rows += 1
            continue
        
        pre_ea_exp_date = standardize_date(pre_ea_exp,in_format="%m/%d/%Y")
        cssm_exp_date = standardize_date(cssm_row['Subscription End Date'])
        if pre_ea_exp_date is None or cssm_exp_date is None:
            logger.warning("Row %d: Invalid date(s). PRE-EA: '%s', CSSM: '%s'. Marking as YELLOW.",
                           excel_row_idx, pre_ea_exp, cssm_row['Subscription End Date'])
            for col in range(1, len(pre_ea.columns) + 1):
                ws.cell(row=excel_row_idx, column=col).fill = YELLOW_FILL
            yellow_rows += 1
            continue

        if pre_ea_exp_date <= cssm_exp_date:
            logger.info("Row %d: Expiration date OK (PRE-EA: %s, CSSM: %s). Marking as GREEN.",
                        excel_row_idx, pre_ea_exp_date, cssm_exp_date)
            for col in range(1, len(pre_ea.columns) + 1):
                ws.cell(row=excel_row_idx, column=col).fill = GREEN_FILL
            green_rows += 1
        else:
            logger.info("Row %d: PRE-EA expiration %s is after CSSM %s.  Marking as YELLOW.",
                        excel_row_idx, pre_ea_exp_date, cssm_exp_date)
            for col in range(1, len(pre_ea.columns) + 1):
                ws.cell(row=excel_row_idx, column=col).fill = YELLOW_FILL
            yellow_rows += 1

    wb.save(out_path)
    logger.info("Finished! Saved comparison result as %s", out_path)
    logger.info("Summary: RED=%d BLUE=%d YELLOW=%d GREEN=%d", red_rows, blue_rows, yellow_rows, green_rows)
    logger.info("Total time: %.2f seconds", (datetime.now() - start_time).total_seconds())

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare PRE-EA and CSSM Excel files and color-code results.")
    parser.add_argument("pre_ea", help="Path to PRE-EA .xlsx file")
    parser.add_argument("cssm", help="Path to CSSM .xlsx file")
    parser.add_argument("-m", "--map", dest="map_path", default=None, help="Path to JSON file with PRE-EA PID to list of CSSM SKUs mapping. If omitted, attempts to use ./sku_map.json if present.")
    return parser.parse_args()


if __name__ == "__main__":
    # Ensure output directory exists and is clean BEFORE configuring logging
    output_dir = os.path.join(os.getcwd(), "output_files")
    try:
        ensure_clean_dir(output_dir)
    except Exception:
        logging.getLogger(__name__).exception("Failed to clean output directory: %s", output_dir)
        sys.exit(1)

    # Configure logging to write inside output directory (fresh log each run)
    setup_logging(log_dir=output_dir, log_filename="compare_excels.log")
    args = parse_args()

    # Print startup message with current date/time and file paths
    print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Starting comparison: pre_ea={args.pre_ea} cssm={args.cssm}")

    try:
        pid_to_skus_map = load_pid_to_skus_map(args.map_path)
        main(args.pre_ea, args.cssm, pid_to_skus_map)
    except Exception:
        logger = logging.getLogger(__name__)
        logger.exception("Unhandled error during comparison")
        sys.exit(1)