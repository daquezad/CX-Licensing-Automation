import os
import sys
import logging
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl import Workbook
# Add parent directory to path for relative imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.fs_utils import ensure_clean_dir
from utils.logging_utils import setup_logging
from utils.mapping_utils import load_pid_to_skus_map, get_valid_sku_matches
from utils.date_utils import standardize_date
from utils.colors import RED_FILL, BLUE_FILL, YELLOW_FILL, GREEN_FILL, PINK_FILL
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelFileComparator:
    def __init__(self, output_dir=None, log_filename="compare_excels.log"):
        self.output_dir = output_dir or os.path.join(os.getcwd(), "output_files")
        ensure_clean_dir(self.output_dir)
        self.logger = setup_logging(log_dir=self.output_dir, log_filename=log_filename)

    def _load_df(self, file_path, sheet_name=None, header=0):
        """Helper method to load a DataFrame from an Excel file."""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
            df.columns = df.columns.map(str.strip)  # Strip whitespace from column names
            return df
        except Exception as e:
            self.logger.error("Error loading file %s: %s", file_path, e)
            raise

    def save_df_with_flag_highlight(self, df_pre_ea, output_path):
        """
        Save df_pre_ea to Excel and highlight rows based on 'Flag' column values.

        Color mapping:
        - GREEN: 🟩 (green fill)
        - RED: 🟥 (red fill)
        - PURPLE: 🟪 (purple fill)
        - YELLOW: 🟨 (yellow fill)
        - BLUE: 🔵 (blue fill)
        - GREY: ⬜ (grey fill)
        """

        # Define fill colors for each flag
        fill_colors = {
            'GREEN': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # light green
            'RED': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),    # light red
            'PURPLE': PatternFill(start_color='D9D2E9', end_color='D9D2E9', fill_type='solid'), # light purple
            'YELLOW': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), # light yellow
            'BLUE': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),   # light blue
            'GREY': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),   # light grey
            '': None  # No fill for empty flags
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "PRE_EA_Report"

        # Write DataFrame headers
        ws.append(list(df_pre_ea.columns))

        # Write DataFrame rows with fill based on 'Flag'
        for idx, row in df_pre_ea.iterrows():
            ws.append(row.tolist())
            flag = row.get('Flag', '')
            fill = fill_colors.get(flag, None)
            if fill:
                # Highlight the entire row (all columns)
                for col_idx in range(1, len(df_pre_ea.columns) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = fill

        # Save the workbook
        wb.save(output_path)
        print(f"DataFrame saved to {output_path} with row highlights based on 'Flag'.")
        
    def find_common_items_in_columns(self,
        df1: pd.DataFrame,
        column1_name: str,
        df2: pd.DataFrame,
        column2_name: str
    ) -> list:

        # Input validation: Check if columns exist in their respective DataFrames
        if column1_name not in df1.columns:
            raise ValueError(f"Column '{column1_name}' not found in DataFrame 1.")
        if column2_name not in df2.columns:
            raise ValueError(f"Column '{column2_name}' not found in DataFrame 2.")
        # Extract unique items from each column and convert to sets for efficient comparison
        set1 = set(df1[column1_name].astype(str).str.strip().unique())
        set2 = set(df2[column2_name].astype(str).str.strip().unique())

        # Find the intersection of the two sets
        common_items_set = set1.intersection(set2)
        # Convert the set of common items to a list and return
        return list(common_items_set)

    def compute_licensing_files(self, pre_ea_path, cssm_path, pid_to_skus_map):
        start_time = datetime.now()
        logger = self.logger
        logger.info("Starting comparison: pre_ea=%s cssm=%s", pre_ea_path, cssm_path)
        start_time = datetime.now()

        # Load CSSM data
        logger.info("Loading CSSM data from %s", cssm_path)
        df_cssm = self._load_df(cssm_path, sheet_name='License Detail', header=5)

        # Load PRE-EA data
        logger.info("Loading PRE-EA data from %s", pre_ea_path)
        df_pre_ea = self._load_df(pre_ea_path,sheet_name='PRE_EA_REPORT', header=0)
        df_pre_ea['Flag'] = ''
        df_pre_ea['Logging Info'] = ''

        # Convert the 'Due_Date' column to datetime objects
        df_pre_ea['Expiration Date'] = pd.to_datetime(df_pre_ea['Expiration Date'])
        today = pd.to_datetime(datetime.today().date())
        mask = df_pre_ea['Expiration Date'] < today
        df_pre_ea.loc[mask, 'Flag'] = 'PURPLE'
        df_pre_ea.loc[mask, 'Logging Info'] = "🟪 Expiration date has already expired." 

        common_codes_dictionary = {}
        df_cssm['Used'] = ''

        # Log common matches: df_pre_ea-> 'ALC Order Number' | df_cssm->'Source Identifier'
        try:
            common_codes_id = self.find_common_items_in_columns(df_pre_ea, 'ALC Order Number', df_cssm, 'Source Identifier')
            self.logger.info(f"Common matches: {len(common_codes_id)}")
        except ValueError as e:
            self.logger.error(f"\nError: {e}")
        # FLAG RED for non-matching ALC Order Numbers
        mask = df_pre_ea['ALC Order Number'].astype(str).str.strip().isin(common_codes_id)
        df_pre_ea.loc[~mask, 'Flag'] = 'RED'
        df_pre_ea.loc[~mask, 'Logging Info'] = "🟥 FLAG RED for non-matching ALC Order Numbers"

        mask = df_cssm['Source Identifier'].astype(str).str.strip().isin(common_codes_id)
        df_cssm.loc[~mask, 'Used'] = 'YES'

        # Normalize pre_ea_migrated_pid_str column to valid sku (usig dictionary)
        if pid_to_skus_map is not None:
            df_pre_ea['Pre EA Migrated Pid'] = df_pre_ea['Pre EA Migrated Pid'].map(pid_to_skus_map).fillna(df_pre_ea['Pre EA Migrated Pid'])

        # FLAG RED for non-matching No SKU (or mapped exception)
        try:
            common_skus = self.find_common_items_in_columns(df_pre_ea, 'Pre EA Migrated Pid', df_cssm, 'SKU')
            self.logger.info(f"Common SKU matches: {len(common_skus)}")
            mask = df_pre_ea['Pre EA Migrated Pid'].astype(str).str.strip().isin(common_skus)
            df_pre_ea.loc[~mask, 'Flag'] = 'RED'
            df_pre_ea.loc[~mask, 'Logging Info'] = "🟥 FLAG RED for non-matching SKU (or mapped exception)"
        except ValueError as e:
            self.logger.error(f"\nError: {e}")
        
        dictionary_filter = {}

        for source_identifier in common_codes_id:
            mask_pre_ea = df_pre_ea['ALC Order Number'].astype(str).str.strip() == str(source_identifier).strip()
            mask_flag = df_pre_ea['Flag'] == ''
            combined_mask_1 = mask_pre_ea & mask_flag
            unique_skus = df_pre_ea[combined_mask_1]['Pre EA Migrated Pid'].astype(str).str.strip().unique()

            # Ensure the dictionary entry is a list before appending
            if source_identifier not in dictionary_filter:
                dictionary_filter[source_identifier] = []
            for sku in unique_skus:
                dictionary_filter[source_identifier].append(sku) # Just call append(), no assignment

        common_codes_dictionary = {}
        for source_id, sku_list in dictionary_filter.items():
            # print(f"Source Identifier: {source_id}")
            if source_id not in common_codes_dictionary:
                common_codes_dictionary[source_id] = {}
            for sku in sku_list:
                # print(f"    Individual SKU: {sku}")
                mask_sku = df_pre_ea['Pre EA Migrated Pid'].astype(str).str.strip() == str(sku).strip()
                mask_flag = df_pre_ea['Flag'] == ''
                mask_pre_ea = df_pre_ea['ALC Order Number'].astype(str).str.strip() == str(source_id).strip()
                combined_mask = mask_pre_ea & mask_sku & mask_flag

                mask_sku_cssm = df_cssm['SKU'].astype(str).str.strip() == sku
                mask_source_id_cssm = df_cssm['Source Identifier'].astype(str).str.strip() == str(source_id).strip()
                combined_mask_cssm = mask_sku_cssm & mask_source_id_cssm
                common_codes_dictionary[source_id][sku] = (df_pre_ea[combined_mask], df_cssm[combined_mask_cssm])
                # print(len(df_pre_ea[combined_mask]),len(df_cssm[combined_mask_cssm]))

        for source_id, sku_dict in common_codes_dictionary.items():
            for sku, (df_pre_ea_subset, df_cssm_subset) in sku_dict.items():
                for pre_ea_index, pre_ea_row in df_pre_ea_subset.iterrows():
                    pre_ea_quantity = pre_ea_row['Quantity']
                    pre_ea_exp_date_str = pre_ea_row['Expiration Date']
                    pre_ea_exp_date = standardize_date(pre_ea_exp_date_str)
                    matched = False

                    for cssm_index, cssm_row in df_cssm_subset.iterrows():
                        cssm_available = cssm_row['Available To Use']
                        if cssm_available == pre_ea_quantity:
                            if pre_ea_exp_date:
                                # Valid expiration date: set green flag
                                print(f"Green: source_id={source_id}, sku={sku}, Quantity={pre_ea_quantity}, Available To Use={cssm_available}, Expiration Date={pre_ea_exp_date_str}")
                                # Example of setting flags in df_pre_ea:
                                df_pre_ea.at[pre_ea_index, 'Flag'] = 'GREEN'
                                df_pre_ea.at[pre_ea_index, 'Logging Info'] = "🟩 FLAG GREEN: Quantity and valid Expiration Date match found."
                            else:
                                # Invalid or empty expiration date: set yellow flag
                                print(f"Yellow: source_id={source_id}, sku={sku}, Quantity={pre_ea_quantity}, Available To Use={cssm_available}, Expiration Date invalid or empty")
                                df_pre_ea.at[pre_ea_index, 'Flag'] = 'YELLOW'
                                df_pre_ea.at[pre_ea_index, 'Logging Info'] = "🟨 FLAG YELLOW: Quantity match found but Expiration Date invalid or empty."
                            matched = True
                            break  # Stop checking CSSM rows once a match is found

                    if not matched:
                        # No matching Available To Use found for this Pre-EA quantity
                        print(f"Blue: source_id={source_id}, sku={sku}, Quantity={pre_ea_quantity} (no matching Available To Use found)")
                        df_pre_ea.at[pre_ea_index, 'Flag'] = 'BLUE'
                        df_pre_ea.at[pre_ea_index, 'Logging Info'] = "🔵 FLAG BLUE: No matching Available To Use found."

        # Loop over common_codes_dictionary again to process BLUE flags
        for source_id, sku_dict in common_codes_dictionary.items():
            for sku, (df_pre_ea_subset, df_cssm_subset) in sku_dict.items():
                # Filter BLUE flagged rows in df_pre_ea_subset
                mask_source = (df_pre_ea['ALC Order Number'].astype(str).str.strip() == str(source_id).strip())
                mask_sku = (df_pre_ea['Pre EA Migrated Pid'].astype(str).str.strip() == str(sku).strip())
                mask_flag_blue = (df_pre_ea['Flag'] == 'BLUE')
                blue_rows = df_pre_ea[mask_source & mask_sku & mask_flag_blue]
                if blue_rows.empty:
                    
                    continue  # No BLUE rows to process for this sku and source_id
                # Sum the Quantity of BLUE rows
                total_blue_quantity = blue_rows['Quantity'].sum()

                # Sum the Quantity in CSSM subset for this sku and source_id
                total_cssm_quantity = df_cssm_subset['Available To Use'].sum()  # or 'Quantity' if that column exists; adjust accordingly
                # Compare totals
                if total_blue_quantity < total_cssm_quantity:
                    
                    # Update all BLUE rows to GREY in prea 
                    for pre_ea_index in blue_rows.index:
                        df_pre_ea.at[pre_ea_index, 'Flag'] = 'GREY'
                        df_pre_ea.at[pre_ea_index, 'Logging Info'] = "⬜ FLAG GREY: Total BLUE Quantity less than CSSM Quantity."
                        # print(f"Grey: source_id={source_id}, sku={sku}, pre_ea_index={pre_ea_index}, Total BLUE Quantity={total_blue_quantity}, CSSM Quantity={total_cssm_quantity}")
        # check point
       

    # counte green flags    
        green_count = df_pre_ea['Flag'].value_counts().get('GREEN', 0)
        print(f"🟩 Number of GREEN rows: {green_count}")
        red_count = df_pre_ea['Flag'].value_counts().get('RED', 0)
        print(f"🟥 Number of RED rows: {red_count}")
        purple_count = df_pre_ea['Flag'].value_counts().get('PURPLE', 0)
        print(f"🟪 Number of PURPLE rows: {purple_count}")
        yellow_count = df_pre_ea['Flag'].value_counts().get('YELLOW', 0)
        print(f"🟨 Number of YELLOW rows: {yellow_count}")
        blue_count = df_pre_ea['Flag'].value_counts().get('BLUE', 0)
        print(f"🔵 Number of BLUE rows: {blue_count}")
        gray_count = df_pre_ea['Flag'].value_counts().get('GREY', 0)
        print(f"⬜ Number of GREY rows: {gray_count}")

        print(len(df_pre_ea))
        # Prepare output workbook

        output_dir = self.output_dir
        base_name = os.path.basename(pre_ea_path)
        out_name = base_name.replace('.xlsx', '_compared.xlsx')
        out_path = os.path.join(output_dir, out_name)

        # Save df_pre_ea to the desired output Excel file
        df_pre_ea.to_excel(out_path, index=False)

        self.save_df_with_flag_highlight(df_pre_ea, out_path)
        return out_path, green_count, red_count, purple_count, yellow_count, blue_count, gray_count, (datetime.now() - start_time).total_seconds()


# Test basic functionality
if __name__ == "__main__":
    print("\n" + "="*60)
    print("BASIC FUNCTIONALITY TEST")
    print("="*60)

    # Define file paths
    pre_ea_path = os.path.join(os.getcwd(), "uploaded_pre_ea.xlsx")
    cssm_path = os.path.join(os.getcwd(), "uploaded_cssm.xlsx")
    pid_to_skus_map = load_pid_to_skus_map(os.path.join(os.getcwd(), "sku_map.json"))

    # print(pid_to_skus_map)
    # Initialize comparator
    comparator = ExcelFileComparator()
    comparator.compute_licensing_files(pre_ea_path, cssm_path, pid_to_skus_map)









