import os
import logging
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from utils.fs_utils import ensure_clean_dir
from utils.logging_utils import setup_logging
from utils.mapping_utils import load_pid_to_skus_map, get_valid_sku_matches
from utils.date_utils import standardize_date
from utils.colors import RED_FILL, BLUE_FILL, YELLOW_FILL, GREEN_FILL, PINK_FILL
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelFileComparator:
    def __init__(self, output_dir=None, log_filename="compare_excels.log"):
        self.output_dir = output_dir or os.path.join(os.getcwd(), "output_files")
        ensure_clean_dir(self.output_dir)
        self.logger = setup_logging(log_dir=self.output_dir, log_filename=log_filename)

    def compare_and_save(self, pre_ea_path, cssm_path, pid_to_skus_map):
        logger = self.logger
        logger.info("Starting comparison: pre_ea=%s cssm=%s", pre_ea_path, cssm_path)
        start_time = datetime.now()

        cssm = pd.read_excel(cssm_path, sheet_name='License Detail', header=5)
        cssm.columns = cssm.columns.map(str.strip)
        cssm['Source Identifier'] = cssm['Source Identifier'].astype(str).str.strip()
        cssm['SKU'] = cssm['SKU'].astype(str).str.strip()

        output_dir = self.output_dir
        base_name = os.path.basename(pre_ea_path)
        out_name = base_name.replace('.xlsx', '_compared.xlsx')
        out_path = os.path.join(output_dir, out_name)
        wb = load_workbook(pre_ea_path)
        wb.save(out_path)
        wb = load_workbook(out_path)
        ws = wb.active

        pre_ea = pd.read_excel(pre_ea_path)
        pre_ea.columns = pre_ea.columns.map(str.strip)

        logger.info("Loaded PRE-EA rows: %d | CSSM rows: %d", len(pre_ea), len(cssm))
        red_rows, blue_rows, yellow_rows, green_rows, pink_rows = 0, 0, 0, 0, 0
        used_cssm_indices = set() # Tracks used CSSM rows to prevent re-matching

        for idx, row in pre_ea.iterrows():
            alc_order_number = row.get('ALC Order Number')
            pre_ea_migrated_pid = row.get('Pre EA Migrated Pid')
            pre_ea_qty = row.get('Quantity')
            pre_ea_exp = row.get('Expiration Date')
            excel_row_idx = idx + 2

            alc_order_number_str = str(alc_order_number).strip()
            pre_ea_migrated_pid_str = str(pre_ea_migrated_pid).strip()
            # --- START OF NEW CONDITION: Check if pre_ea_exp is later than today ---

            pre_ea_exp_date_for_check = standardize_date(pre_ea_exp, in_format="%m/%d/%Y")
            today = datetime.today().date() # Get today's date for comparison

            if pre_ea_exp_date_for_check and pre_ea_exp_date_for_check < today: # Changed from > to <
                self.logger.info("Row %d: PRE-EA Expiration Date '%s' is earlier than today. Marking as 🟪 PURPLE.", excel_row_idx, pre_ea_exp_date_for_check)
                for col in range(1, len(pre_ea.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = PINK_FILL 
                pink_rows += 1
                continue 
            cssm_matches = cssm[cssm['Source Identifier'] == alc_order_number_str]
            if cssm_matches.empty:
                logger.info("Row %d: ALC Order Number '%s' NOT found in CSSM. Marking as 🟥 RED.", excel_row_idx, alc_order_number_str)
                for col in range(1, len(pre_ea.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = RED_FILL
                red_rows += 1
                continue
            sku_match = get_valid_sku_matches(cssm_matches, pre_ea_migrated_pid_str, pid_to_skus_map)

            if sku_match.empty:
                logger.info("Row %d: No SKU '%s' (or mapped exception) for ALC Order Number '%s' in CSSM. Marking as 🟥 RED.", excel_row_idx, pre_ea_migrated_pid_str, alc_order_number_str)
                for col in range(1, len(pre_ea.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = RED_FILL
                red_rows += 1
                continue

            quantity_match_found = False
            matched_cssm_row = None

            # Iterate through each potential SKU match from the CSSM data
            for cssm_index, cssm_row_iter in sku_match.iterrows():
                # Check if this specific CSSM row has already been used for a previous match
                if cssm_index in used_cssm_indices:
                    continue  # This CSSM entry is already matched, skip to the next one

                # Safely get and convert the quantity from the CSSM row
                try:
                    cssm_qty = int(cssm_row_iter['Available To Use'])
                except (ValueError, TypeError):
                    # If conversion fails, this row cannot be a valid quantity match
                    continue

                # Compare quantities
                if cssm_qty == pre_ea_qty:
                    # We found a valid, unused match!
                    quantity_match_found = True
                    # Mark this CSSM row's index as used so it can't be matched again
                    used_cssm_indices.add(cssm_index)
                    # Store the matched row for the subsequent date comparison
                    matched_cssm_row = cssm_row_iter
                    # No need to check other potential SKU matches for this pre_ea row
                    break
            
            # After checking all potential SKU matches, evaluate if one was found
            if not quantity_match_found:
                logger.info("Row %d: Quantity mismatch (PRE-EA: %s, CSSM: No available matching quantity). Marking as 🟦 BLUE.", excel_row_idx, pre_ea_qty)
                for col in range(1, len(pre_ea.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = BLUE_FILL
                blue_rows += 1
                continue
            else:
                # A match was found, so we now use the stored 'matched_cssm_row'
                cssm_row = matched_cssm_row

            pre_ea_exp_date = standardize_date(pre_ea_exp, in_format="%m/%d/%Y")
            cssm_exp_date = standardize_date(cssm_row['Subscription End Date'])
            if pre_ea_exp_date is None or cssm_exp_date is None:
                logger.warning("Row %d: Invalid date(s). PRE-EA: '%s', CSSM: '%s'. Marking as 🟨 YELLOW.", excel_row_idx, pre_ea_exp, cssm_row['Subscription End Date'])
                for col in range(1, len(pre_ea.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = YELLOW_FILL
                yellow_rows += 1
                continue

            if pre_ea_exp_date <= cssm_exp_date:
                logger.info("Row %d: Expiration date OK (PRE-EA: %s, CSSM: %s). Marking as 🟩 GREEN.", excel_row_idx, pre_ea_exp_date, cssm_exp_date)
                for col in range(1, len(pre_ea.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = GREEN_FILL
                green_rows += 1
            else:
                logger.info("Row %d: PRE-EA expiration %s is after CSSM %s.  Marking as 🟨 YELLOW.", excel_row_idx, pre_ea_exp_date, cssm_exp_date)
                for col in range(1, len(pre_ea.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = YELLOW_FILL
                yellow_rows += 1

        wb.save(out_path)
        logger.info(f"✅ Finished! Saved comparison result as {out_path}")
        logger.info(f"Summary: 🟥 RED={red_rows} 🟦 BLUE={blue_rows} 🟨 YELLOW={yellow_rows} 🟩 GREEN={green_rows} 🟪 PURPLE={pink_rows}")
        logger.info(f"⏱️ Total time: {(datetime.now() - start_time).total_seconds():.2f} seconds")
        for handler in logger.handlers:
            handler.flush()
        return out_path, red_rows, blue_rows, yellow_rows, green_rows, pink_rows, (datetime.now() - start_time).total_seconds()
