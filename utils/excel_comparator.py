import os
import io
import logging
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from utils.fs_utils import ensure_clean_dir
from utils.logging_utils import setup_logging
from utils.mapping_utils import load_pid_to_skus_map, get_valid_sku_matches
from utils.date_utils import standardize_date
from utils.colors import RED_FILL, BLUE_FILL, YELLOW_FILL, GREEN_FILL

class ExcelComparator:
    def __init__(self, output_dir=None, log_filename="compare_excels.log"):
        self.output_dir = output_dir or os.path.join(os.getcwd(), "output_files")
        ensure_clean_dir(self.output_dir)
        self.logger = setup_logging(log_dir=self.output_dir, log_filename=log_filename)

    def compare_excels_in_memory(self, pre_ea_bytes: bytes, cssm_bytes: bytes, pid_to_skus_map: dict[str, list[str]]):
        start_time = datetime.now()
        pre_ea_df = pd.read_excel(io.BytesIO(pre_ea_bytes))
        pre_ea_df.columns = pre_ea_df.columns.map(str.strip)
        cssm_df = pd.read_excel(io.BytesIO(cssm_bytes), sheet_name='License Detail', header=5)
        cssm_df.columns = cssm_df.columns.map(str.strip)
        cssm_df['Source Identifier'] = cssm_df['Source Identifier'].astype(str).str.strip()
        cssm_df['SKU'] = cssm_df['SKU'].astype(str).str.strip()

        wb = load_workbook(io.BytesIO(pre_ea_bytes))
        ws = wb.active

        self.logger.info("Loaded PRE-EA rows: %d | CSSM rows: %d", len(pre_ea_df), len(cssm_df))
        red_rows, blue_rows, yellow_rows, green_rows = 0, 0, 0, 0
        for idx, row in pre_ea_df.iterrows():
            alc_order_number = row.get('ALC Order Number')
            pre_ea_migrated_pid = row.get('Pre EA Migrated Pid')
            pre_ea_qty = row.get('Quantity')
            pre_ea_exp = row.get('Expiration Date')
            excel_row_idx = idx + 2

            alc_order_number_str = str(alc_order_number).strip()
            pre_ea_migrated_pid_str = str(pre_ea_migrated_pid).strip()

            cssm_matches = cssm_df[cssm_df['Source Identifier'] == alc_order_number_str]
            if cssm_matches.empty:
                self.logger.info("Row %d: ALC Order Number '%s' NOT found in CSSM. Marking as 🟥 RED.", excel_row_idx, alc_order_number_str)
                for col in range(1, len(pre_ea_df.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = RED_FILL
                red_rows += 1
                continue

            sku_match = get_valid_sku_matches(cssm_matches, pre_ea_migrated_pid_str, pid_to_skus_map)
            if sku_match.empty:
                self.logger.info("Row %d: No SKU '%s' (or mapped exception) for ALC Order Number '%s' in CSSM. Marking as 🟥 RED.", excel_row_idx, pre_ea_migrated_pid_str, alc_order_number_str)
                for col in range(1, len(pre_ea_df.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = RED_FILL
                red_rows += 1
                continue

            cssm_row = sku_match.iloc[0]
            cssm_qty = cssm_row['Available To Use']
            try:
                cssm_qty = int(cssm_qty)
            except Exception:
                cssm_qty = None

            if cssm_qty != pre_ea_qty:
                self.logger.info("Row %d: Quantity mismatch (PRE-EA: %s, CSSM: %s). Marking as 🟦 BLUE.", excel_row_idx, pre_ea_qty, cssm_qty)
                for col in range(1, len(pre_ea_df.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = BLUE_FILL
                blue_rows += 1
                continue

            pre_ea_exp_date = standardize_date(pre_ea_exp, in_format="%m/%d/%Y")
            cssm_exp_date = standardize_date(cssm_row['Subscription End Date'])
            if pre_ea_exp_date is None or cssm_exp_date is None:
                self.logger.warning("Row %d: Invalid date(s). PRE-EA: '%s', CSSM: '%s'. Marking as 🟨 YELLOW.", excel_row_idx, pre_ea_exp, cssm_row['Subscription End Date'])
                for col in range(1, len(pre_ea_df.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = YELLOW_FILL
                yellow_rows += 1
                continue

            if pre_ea_exp_date <= cssm_exp_date:
                self.logger.info("Row %d: Expiration date OK (PRE-EA: %s, CSSM: %s). Marking as 🟩 GREEN.", excel_row_idx, pre_ea_exp_date, cssm_exp_date)
                for col in range(1, len(pre_ea_df.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = GREEN_FILL
                green_rows += 1
            else:
                self.logger.info("Row %d: PRE-EA expiration %s is after CSSM %s.  Marking as 🟨 YELLOW.", excel_row_idx, pre_ea_exp_date, cssm_exp_date)
                for col in range(1, len(pre_ea_df.columns) + 1):
                    ws.cell(row=excel_row_idx, column=col).fill = YELLOW_FILL
                yellow_rows += 1

        self.logger.info("Summary: RED=%d BLUE=%d YELLOW=%d GREEN=%d", red_rows, blue_rows, yellow_rows, green_rows)
        self.logger.info("Total time: %.2f seconds", (datetime.now() - start_time).total_seconds())
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        for handler in self.logger.handlers:
            handler.flush()
        return out_buf
